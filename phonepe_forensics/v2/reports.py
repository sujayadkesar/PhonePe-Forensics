"""CSV/XLSX writers. The HTML report is built in static_export.py."""
from __future__ import annotations

import csv
import json
from dataclasses import asdict
from pathlib import Path
from typing import Iterable

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

from .core import ts_ms_to_naive_ist

PAYMENT_COLUMNS = [
    "primary_id",
    "global_id",
    "data_source",
    "entity_type",
    "classification",
    "merchant_subtype",
    "direction",
    "amount_inr",
    "amount_paise",
    "state",
    "is_failed_chat_only",
    "failure_reason",
    "is_refund",
    "timestamp_ms",
    "datetime_ist",
    "datetime_utc",
    "counterparty_name",
    "counterparty_verified_name",
    "counterparty_cbs_name",
    "counterparty_phone",
    "counterparty_vpa",
    "counterparty_full_vpa",
    "receiver_vpa",
    "sender_vpa",
    "sender_app_label",
    "sender_app_source",
    "receiver_app_label",
    "receiver_app_source",
    "sender_on_phonepe",
    "receiver_on_phonepe",
    "bank_id",
    "ifsc",
    "bank_name",
    "mcc",
    "service_type",
    "merchant_id",
    "merchant_type",
    "merchant_genre",
    "merchant_identifier_type",
    "first_party_merchant",
    "payment_initiation",
    "upi_initiation_mode",
    "transfer_mode",
    "is_qr_scan",
    "is_intent",
    "intent_caller_url",
    "note",
    "utr",
]


def _payment_as_row(p) -> dict:
    d = asdict(p)
    d.pop("decoded_blob", None)
    d.pop("provenance", None)
    d.pop("sender_app_icon_id", None)
    d.pop("receiver_app_icon_id", None)
    return d


def write_payments_csv(payments: Iterable, out_path: Path) -> int:
    rows = [_payment_as_row(p) for p in payments]
    with open(out_path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=PAYMENT_COLUMNS, extrasaction="ignore")
        w.writeheader()
        for r in rows:
            w.writerow(r)
    return len(rows)


def write_payments_xlsx(payments: Iterable, out_path: Path) -> int:
    if not HAS_OPENPYXL:
        return 0
    rows = [_payment_as_row(p) for p in payments]
    wb = Workbook()
    ws = wb.active
    ws.title = "Payments"
    ws.append(PAYMENT_COLUMNS)
    header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    dt_col_idx = PAYMENT_COLUMNS.index("datetime_ist") + 1
    for r in rows:
        out_row = []
        for k in PAYMENT_COLUMNS:
            v = r.get(k)
            if k == "datetime_ist":
                v = ts_ms_to_naive_ist(r.get("timestamp_ms"))
            out_row.append(v)
        ws.append(out_row)
    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row=row, column=dt_col_idx)
        cell.number_format = "YYYY-MM-DD HH:MM:SS"
    # auto column widths (bounded)
    for col in ws.columns:
        letter = col[0].column_letter
        width = min(max((len(str(c.value)) for c in col if c.value), default=10) + 2, 40)
        ws.column_dimensions[letter].width = width
    wb.save(out_path)
    return len(rows)


def write_mandates_csv(mandates: list[dict], out_path: Path) -> int:
    if not mandates:
        return 0
    cols = [
        "primary_id",
        "global_id",
        "entity_type",
        "state",
        "timestamp_ms",
        "datetime_ist",
        "name",
        "amount_inr",
        "amount_paise",
    ]
    with open(out_path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=cols, extrasaction="ignore")
        w.writeheader()
        for m in mandates:
            w.writerow({c: m.get(c) for c in cols})
    return len(mandates)


def write_failed_payments_csv(payments: Iterable, out_path: Path) -> int:
    rows = [
        _payment_as_row(p)
        for p in payments
        if p.state == "FAILED"
    ]
    if not rows:
        return 0
    with open(out_path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=PAYMENT_COLUMNS, extrasaction="ignore")
        w.writeheader()
        for r in rows:
            w.writerow(r)
    return len(rows)


def write_refunds_csv(payments: Iterable, mandates: list[dict], out_path: Path) -> int:
    payment_rows = [_payment_as_row(p) for p in payments if p.is_refund]
    mandate_rows = [m for m in mandates if m.get("is_merchant_refund")]
    if not payment_rows and not mandate_rows:
        return 0
    cols = [
        "primary_id",
        "kind",
        "entity_type",
        "state",
        "amount_inr",
        "datetime_ist",
        "counterparty_name",
        "merchant_name",
    ]
    with open(out_path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=cols, extrasaction="ignore")
        w.writeheader()
        for r in payment_rows:
            w.writerow(
                {
                    "primary_id": r["primary_id"],
                    "kind": "payment_refund",
                    "entity_type": r["entity_type"],
                    "state": r["state"],
                    "amount_inr": r["amount_inr"],
                    "datetime_ist": r["datetime_ist"],
                    "counterparty_name": r["counterparty_name"],
                    "merchant_name": r["counterparty_name"],
                }
            )
        for m in mandate_rows:
            w.writerow(
                {
                    "primary_id": m["primary_id"],
                    "kind": "request_from_merchant",
                    "entity_type": m["entity_type"],
                    "state": m["state"],
                    "amount_inr": m["amount_inr"],
                    "datetime_ist": m["datetime_ist"],
                    "counterparty_name": m["name"],
                    "merchant_name": m["name"],
                }
            )
    return len(payment_rows) + len(mandate_rows)


def write_raw_records_jsonl(payments: Iterable, mandates: list[dict], out_path: Path) -> int:
    n = 0
    with open(out_path, "w", encoding="utf-8") as f:
        for p in payments:
            obj = {
                "kind": "payment",
                "primary_id": p.primary_id,
                "provenance": p.provenance,
                "decoded": p.decoded_blob,
            }
            f.write(json.dumps(obj, ensure_ascii=False, default=str) + "\n")
            n += 1
        for m in mandates:
            obj = {
                "kind": "mandate_or_request",
                "primary_id": m.get("primary_id"),
                "provenance": m.get("provenance"),
                "decoded": m.get("decoded_blob"),
            }
            f.write(json.dumps(obj, ensure_ascii=False, default=str) + "\n")
            n += 1
    return n
